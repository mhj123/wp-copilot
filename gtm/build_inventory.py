#!/usr/bin/env python3
"""Build the Work Copilot feature inventory workbook for GTM scope decisions."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT = "Arial"

# ---------------------------------------------------------------- styling ---
H_FILL = PatternFill("solid", fgColor="1F3864")
H_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT, size=14, bold=True, color="1F3864")
SUB_FONT = Font(name=FONT, size=10, italic=True, color="595959")
BODY = Font(name=FONT, size=10)
BODY_B = Font(name=FONT, size=10, bold=True)
INPUT_FONT = Font(name=FONT, size=10, color="0000FF")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
GROUP_FILL = PatternFill("solid", fgColor="D9E2F3")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")

REC_FILLS = {
    "Keep - core":               PatternFill("solid", fgColor="C6EFCE"),
    "Keep - flag off by default":PatternFill("solid", fgColor="FFEB9C"),
    "Fix before launch":         PatternFill("solid", fgColor="F8CBAD"),
    "Extract to add-on":         PatternFill("solid", fgColor="BDD7EE"),
    "Move to backup repo":       PatternFill("solid", fgColor="D9D9D9"),
    "Delete":                    PatternFill("solid", fgColor="FFC7CE"),
}

RECOMMENDATIONS = list(REC_FILLS.keys())

# ------------------------------------------------------------------ data ---
# (bundle, area, feature, what it does, code location, usp, effort_to_keep,
#  launch_risk, recommendation, rationale)
ROWS = [
    # ============================ CORE: CONTENT MODEL ============================
    ("Core plugin", "Content model", "Items as native WP posts",
     "Atomic unit of the system - one task/idea/note per post.",
     "wp-copilot/includes/class-post-types.php", "U1, U4", "Low", "None",
     "Keep - core", "The atomic-item USP. Non-negotiable."),
    ("Core plugin", "Content model", "Pages as contexts",
     "Native WP pages form the folder-like hierarchy for areas and projects.",
     "wp-copilot/includes/class-taxonomy-sync.php", "U3, U5", "Low", "None",
     "Keep - core", "The organisation USP. Non-negotiable."),
    ("Core plugin", "Content model", "Headings CPT (wcp_heading)",
     "Sub-sections inside a page that group items without polluting the page tree.",
     "wp-copilot/includes/class-post-types.php", "U3, U4", "Low", "None",
     "Keep - core", "Needed for page structure to be usable beyond a flat list."),
    ("Core plugin", "Content model", "wcp_context taxonomy + auto-sync",
     "Mirrors the page/heading tree into a hierarchical taxonomy; how items attach to structure.",
     "wp-copilot/includes/class-taxonomy-sync.php", "U3, U4", "Med", "Sync can drift on bulk edits - needs a repair/verify pass",
     "Keep - core", "Load-bearing for everything. Verify sync integrity before launch."),
    ("Core plugin", "Content model", "Multi-context association",
     "One item can belong to several pages/headings at once.",
     "wp-copilot/includes/class-rest-api.php", "U4", "Low", "None",
     "Keep - core", "This is what makes canonical references possible."),
    ("Core plugin", "Content model", "item_type taxonomy (task/info/learning/spec)",
     "Classifies what kind of atom each item is.",
     "wp-copilot/includes/class-taxonomies.php", "U4", "Low", "'spec' type is thinly used",
     "Keep - core", "Keep task/info/learning. Consider dropping 'spec' - it duplicates 'info' for most users."),
    ("Core plugin", "Content model", "priority taxonomy (critical/high/med/low)",
     "Priority ranking on items.",
     "wp-copilot/includes/class-taxonomies.php", "U4", "Low", "None",
     "Keep - core", "Expected baseline for a work tool."),
    ("Core plugin", "Content model", "task_status taxonomy (to-do/in-progress/done)",
     "Workflow state for task items.",
     "wp-copilot/includes/class-taxonomies.php", "U4", "Low", "None",
     "Keep - core", "Expected baseline."),
    ("Core plugin", "Content model", "spec_status taxonomy (draft/review/final)",
     "Workflow state for spec items only.",
     "wp-copilot/includes/class-taxonomies.php", "U4", "Low", "Extra taxonomy for a rarely used item type",
     "Keep - flag off by default", "Adds a UI dropdown most users never touch. Ship hidden."),
    ("Core plugin", "Content model", "Pinned taxonomy",
     "Floats an item to the top of its page.",
     "wp-copilot/includes/class-taxonomies.php", "U4", "Low", "None",
     "Keep - core", "Cheap, high-utility."),
    ("Core plugin", "Content model", "Freeform tags (native post_tag)",
     "Standard WP tags on items, with tag archive pages.",
     "wp-copilot/includes/class-taxonomies.php", "U4", "Low", "None",
     "Keep - core", "Tagging is explicitly part of the atomic-item pitch."),
    ("Core plugin", "Content model", "Nested sub-items (post_parent outliner)",
     "Items nest to arbitrary depth and render as an indented outline.",
     "work-copilot-theme/functions.php (render_item_tree)", "U4", "Med", "Deep nesting + context filters interact subtly",
     "Keep - core", "Needed for action plans to land as real structure."),
    ("Core plugin", "Content model", "Subtasks (JSON in postmeta)",
     "Lightweight checklist inside a single item.",
     "wp-copilot/includes/class-rest-api.php", "U4, U6", "Low", "JSON-in-meta is not queryable",
     "Keep - core", "Target for AI action plans. Keep, but do not extend it further - use nested items instead."),
    ("Core plugin", "Content model", "Due dates",
     "Date meta on items, drives the homepage 'this week' view.",
     "wp-copilot/includes/class-rest-api.php", "U4", "Low", "No timezone handling review yet",
     "Keep - core", "Baseline work-management expectation."),
    ("Core plugin", "Content model", "Source URL meta",
     "Optional origin link on an item.",
     "wp-copilot/includes/class-rest-api.php", "U4", "Low", "None",
     "Keep - core", "Trivial to keep, needed by importers."),
    ("Core plugin", "Content model", "Creator mark (manual/copilot/hermes)",
     "Badges an item by who created it, plus an admin list filter/column.",
     "wp-copilot/admin/class-admin.php", "U6", "Low", "'hermes' value leaks the delegation add-on into core",
     "Keep - core", "Provenance is central to the co-authoring loop. Rename 'hermes' to a generic 'agent'."),
    ("Core plugin", "Content model", "Manual reorder (menu_order)",
     "User-defined ordering of items within a context.",
     "wp-copilot/includes/class-rest-api.php", "U4", "Low", "None",
     "Keep - core", "Expected in an outliner."),

    # ============================ CORE: AI ENGINE ============================
    ("Core plugin", "AI engine", "Anthropic Claude API client",
     "All AI generation and chat calls.",
     "wp-copilot/includes/class-ai-client.php", "U1", "Med", "No retry/backoff or rate-limit handling; error surfacing is thin",
     "Fix before launch", "BYO-key is the headline USP - failures must be legible to a non-technical user."),
    ("Core plugin", "AI engine", "Global AI on/off toggle",
     "Plugin is fully usable with AI disabled.",
     "wp-copilot/admin/class-settings.php", "U1", "Low", "None",
     "Keep - core", "Also the WP.org compliance story: no key, no external calls."),
    ("Core plugin", "AI engine", "Model selection + per-session override",
     "Site default model, plus a Haiku/Sonnet/Opus picker in the widget.",
     "work-copilot-theme/template-parts/ai-widget.php", "U1", "Low", "Model IDs are hardcoded in the template - will rot",
     "Fix before launch", "Move the model list to a filterable config array so it survives model releases."),
    ("Core plugin", "AI engine", "Thinking budget selector",
     "Off/low/med/high extended thinking, Opus only.",
     "work-copilot-theme/template-parts/ai-widget.php", "U1", "Low", "Power-user control in the primary UI",
     "Keep - flag off by default", "Confusing on first run. Expose in settings for advanced users."),
    ("Core plugin", "AI engine", "Global mission",
     "Site-wide brief describing what the user is working on.",
     "wp-copilot/admin/class-settings.php", "U5", "Low", "None",
     "Keep - core", "The AGENTS.md-equivalent USP."),
    ("Core plugin", "AI engine", "Global instructions",
     "Behavioural constraints applied to every prompt.",
     "wp-copilot/admin/class-settings.php", "U5", "Low", "Overlaps conceptually with global mission",
     "Keep - core", "Merge the two fields in the UI if onboarding testing shows confusion."),
    ("Core plugin", "AI engine", "Page-level mission metabox",
     "Per-page brief the assistant works against on that page.",
     "wp-copilot/admin/class-page-mission-metabox.php", "U5", "Low", "Only editable from wp-admin, not the front end",
     "Fix before launch", "Core USP but hidden in wp-admin - surface it in the workspace UI."),
    ("Core plugin", "AI engine", "Mission loader / inheritance",
     "Resolves which mission applies on a given page.",
     "wp-copilot/includes/class-mission-loader.php", "U5", "Med", "Inheritance rules are undocumented for users",
     "Keep - core", "Document the precedence order in the onboarding guide."),
    ("Core plugin", "AI engine", "Context builder (context packs)",
     "Assembles the hierarchical page/heading/item context sent to the model.",
     "wp-copilot/includes/class-context-builder.php", "U5", "High", "No visible token budget or truncation feedback to the user",
     "Fix before launch", "The single most important file for output quality. Add a 'what was sent' disclosure."),
    ("Core plugin", "AI engine", "Prompt builder (layered prompts)",
     "Assembles system + user prompts from mission, instructions and context.",
     "wp-copilot/includes/class-prompt-builder.php", "U5", "High", "None",
     "Keep - core", "Core differentiator vs a generic chat app."),
    ("Core plugin", "AI engine", "Context scope selector (page / corpus / select pages)",
     "Chooses what context the assistant sees for this turn.",
     "work-copilot-theme/template-parts/ai-widget.php", "U5", "Low", "'Corpus (RAG)' silently requires an OpenAI key",
     "Fix before launch", "Explicit USP. Must degrade gracefully and explain why RAG is unavailable."),
    ("Core plugin", "AI engine", "Conversation persistence",
     "Per-page conversation sessions with the last 10 turns replayed.",
     "wp-copilot/includes/class-conversations-manager.php", "U3", "Med", "No pruning/retention policy on the messages table",
     "Keep - core", "This IS the 'organised AI chats' USP. Add retention controls."),
    ("Core plugin", "AI engine", "AI audit log (table + admin viewer)",
     "Records model, prompt, input, output and the accept/dismiss decision for every call.",
     "wp-copilot/includes/class-ai-logger.php", "U1", "Med", "Table grows unbounded; stores full prompts (privacy)",
     "Fix before launch", "Great trust story and a WP.org asset - needs retention + a purge button."),
    ("Core plugin", "AI engine", "Proposal / accept-dismiss workflow",
     "Nothing the AI produces is written until the user accepts it.",
     "wp-copilot/includes/class-ai-actions.php (execute_proposal)", "U6", "High", "None",
     "Keep - core", "The safety guarantee the whole pitch rests on."),
    ("Core plugin", "AI engine", "Auto-route (intent detection)",
     "Picks which AI action to run from a free-form prompt.",
     "wp-copilot/includes/class-ai-actions.php (auto_route)", "U5", "Med", "Misroutes are confusing and hard to debug",
     "Keep - flag off by default", "Nice when right, baffling when wrong. Ship with explicit action selection as default."),

    # ============================ CORE: PAGE AI ACTIONS ============================
    ("Core plugin", "AI actions (page)", "Chat / Q&A",
     "Free-form conversation grounded in the current page's context.",
     "class-ai-actions.php: chat_qa()", "U3, U5", "Med", "None",
     "Keep - core", "The default interaction. Must be flawless."),
    ("Core plugin", "AI actions (page)", "Onboard",
     "Greets the user, summarises the page, suggests next steps.",
     "class-ai-actions.php: onboard()", "U5", "Low", "None",
     "Keep - core", "Doubles as product onboarding - lean on this in the guide."),
    ("Core plugin", "AI actions (page)", "Generate structure",
     "Proposes headings plus placed items for a page.",
     "class-ai-actions.php: generate_structure()", "U6", "Med", "None",
     "Keep - core", "Strongest first-run 'wow' moment. Feature it in demo content."),
    ("Core plugin", "AI actions (page)", "Generate items",
     "Produces new items from a prompt.",
     "class-ai-actions.php: generate_items()", "U6", "Med", "None",
     "Keep - core", "Bread and butter."),
    ("Core plugin", "AI actions (page)", "Generate headings",
     "Produces headings only.",
     "class-ai-actions.php: generate_headings()", "U6", "Low", "Overlaps 'generate structure'",
     "Keep - flag off by default", "Consolidate into generate structure to cut menu clutter."),
    ("Core plugin", "AI actions (page)", "Generate pages (sub-pages)",
     "Suggests child pages to create.",
     "class-ai-actions.php: generate_pages()", "U3, U6", "Med", "None",
     "Keep - core", "Directly serves the 'build your structure' promise."),
    ("Core plugin", "AI actions (page)", "Edit items (bulk AI edit)",
     "Rewrites titles/content across multiple items in one proposal.",
     "class-ai-actions.php: edit_items()", "U6", "High", "Known JSON parse/truncation issues (see changelog 1.2.2)",
     "Fix before launch", "High blast radius. Needs hardened parsing and a hard cap on batch size."),
    ("Core plugin", "AI actions (page)", "Rewrite page content",
     "Replaces the page's content block.",
     "class-ai-actions.php: rewrite_page_content()", "U6", "Med", "Destructive - no visible undo path",
     "Fix before launch", "Snapshot to a WP revision before applying."),
    ("Core plugin", "AI actions (page)", "Append to page content",
     "Adds new content to the end of the page.",
     "class-ai-actions.php: append_page_content()", "U6", "Low", "None",
     "Keep - core", "Safe counterpart to rewrite."),
    ("Core plugin", "AI actions (page)", "Fetch posts (interpret + execute)",
     "Two-step query builder that pulls existing items matching a description.",
     "class-ai-actions.php: fetch_posts_*()", "U5", "High", "Overlaps search, dynamic listings and chat",
     "Keep - flag off by default", "Three ways to do the same job. Pick one for launch."),
    ("Core plugin", "AI actions (page)", "Fetch structure",
     "Answers questions about the existing page/heading tree.",
     "class-ai-actions.php: fetch_structure_chat()", "U5", "Med", "Overlaps chat",
     "Keep - flag off by default", "Chat with corpus context already covers this."),
    ("Core plugin", "AI actions (page)", "Coaching dialogue",
     "Multi-turn coaching conversation over the page's learnings.",
     "class-ai-actions.php: coaching_dialogue()", "U3, U5", "Med", "Distinct mode with no obvious entry point",
     "Keep - flag off by default", "Strong concept, weak discoverability. Revisit post-launch as a prompt preset."),
    ("Core plugin", "AI actions (page)", "Generate single item",
     "Produces exactly one item.",
     "class-ai-actions.php: generate_single_item()", "U6", "Low", "Duplicate of generate items with count=1",
     "Delete", "Collapse into generate_items with an item count parameter."),
    ("Core plugin", "AI actions (page)", "Taxonomy outline",
     "Summarises the taxonomy/structure as an outline.",
     "class-ai-actions.php: taxonomy_outline()", "-", "Low", "Debug-grade utility exposed as a user feature",
     "Move to backup repo", "Developer tool, not a user feature."),
    ("Core plugin", "AI actions (page)", "Mission priorities",
     "Ranks what to work on next against the mission.",
     "class-ai-actions.php: mission_priorities()", "U5", "Med", "No entry point in the primary UI",
     "Keep - flag off by default", "Good idea, unfinished surface. Post-launch."),
    ("Core plugin", "AI actions (page)", "Weekly summary / activity summary",
     "Summarises what changed in the last week; shown on the homepage.",
     "class-ai-actions.php: weekly_summary()", "U6", "Med", "None",
     "Keep - core", "Recurring-value hook that pulls people back in."),
    ("Core plugin", "AI actions (page)", "Goal planning (2-step)",
     "Turns a description into a goal with a heading and action items.",
     "class-ai-actions.php: plan_goal() + /goals/create", "U6", "Med", "Introduces a 'goal' concept not in the data model",
     "Keep - flag off by default", "Conceptual sprawl - a goal is just a page or heading. Simplify or hide."),
    ("Core plugin", "AI actions (page)", "Summarise page",
     "One-shot summary of a page and its items.",
     "class-ai-actions.php: summarize_page()", "U5", "Low", "None",
     "Keep - core", "Cheap and obviously useful."),
    ("Core plugin", "AI actions (page)", "Expand draft in wp-admin editor",
     "AI metabox on the classic editor that expands a draft.",
     "wp-copilot/admin/class-admin.php: render_editor_ai_meta_box()", "-", "Low", "Second AI surface in a second UI to maintain",
     "Move to backup repo", "Splits the product story. Everything should happen in the workspace UI (U7)."),
    ("Core plugin", "AI actions (page)", "Memory extraction",
     "Pulls 1-3 learnings out of a conversation and proposes them as items.",
     "wp-copilot/includes/class-memory-manager.php", "U6", "Med", "Requires a manually designated Memories page",
     "Keep - core", "This is the learning-loop USP made concrete. Auto-provision the Memories page."),
    ("Core plugin", "AI actions (page)", "Save chat message as item",
     "Turns any assistant message into a stored item.",
     "work-copilot-theme/assets/js/ai-widget.js", "U4, U6", "Low", "None",
     "Keep - core", "The bridge from 'wall of text' to atomic items. Central to the pitch."),

    # ============================ CORE: ITEM AI ACTIONS ============================
    ("Core plugin", "AI actions (item)", "Action plan",
     "Breaks an item into 4-7 ordered steps.",
     "class-rest-api.php: case 'action_plan'", "U6", "Med", "None",
     "Keep - core", "Most-used item action."),
    ("Core plugin", "AI actions (item)", "Action plan from context (RAG-grounded)",
     "Finds relevant SOP/reference pages first, then plans against them.",
     "class-rest-api.php: case 'action_plan_from_context'", "U5, U6", "High", "Requires embeddings; multi-step UI",
     "Keep - core", "The clearest proof of 'works against your own protocols'. Flagship demo."),
    ("Core plugin", "AI actions (item)", "Improve phrasing",
     "Rewrites an item's title and content for clarity.",
     "class-rest-api.php: case 'improve_phrasing'", "U4", "Low", "None",
     "Keep - core", "Cheap and popular."),
    ("Core plugin", "AI actions (item)", "Freeform instruction",
     "Apply any user instruction to an item.",
     "class-rest-api.php: case 'freeform'", "U6", "Low", "None",
     "Keep - core", "Escape hatch that covers the long tail."),
    ("Core plugin", "AI actions (item)", "Suggest subtasks",
     "Proposes 3-6 concrete subtasks.",
     "class-rest-api.php: case 'suggest_subtasks'", "U6", "Low", "Overlaps action plan",
     "Keep - flag off by default", "Action plan already offers 'accept as subtasks'."),
    ("Core plugin", "AI actions (item)", "Auto-associate contexts",
     "Suggests which pages/headings an item belongs to.",
     "class-rest-api.php: case 'suggest_contexts'", "U4", "Med", "None",
     "Keep - core", "Directly reduces the filing burden - a real adoption unlock."),
    ("Core plugin", "AI actions (item)", "Convert to goal",
     "Pre-fills goal creation from an item.",
     "class-rest-api.php: case 'to_goal'", "-", "Low", "Depends on the goal concept",
     "Keep - flag off by default", "Ships or hides with the goals feature."),
    ("Core plugin", "AI actions (item)", "Accept plan as subtasks or nested items",
     "Choose how generated steps land in the corpus.",
     "work-copilot-theme/assets/js/theme.js", "U4, U6", "Med", "None",
     "Keep - core", "The co-authoring mechanic in one interaction."),

    # ============================ CORE: SEMANTIC SEARCH ============================
    ("Core plugin", "Semantic search", "OpenAI embeddings client",
     "Generates text-embedding-3-small vectors for items/pages/headings.",
     "wp-copilot/includes/class-embeddings-client.php", "U5", "Med", "Second required API key; provider hardcoded",
     "Fix before launch", "Two keys is real onboarding friction. Abstract the provider and consider a keyless local fallback."),
    ("Core plugin", "Semantic search", "Auto-embed on save",
     "Re-embeds content whenever it changes.",
     "wp-copilot/includes/class-embeddings-manager.php", "U5", "Med", "Synchronous cost/latency on every save",
     "Fix before launch", "Move to a queued background job - this will bite on large corpora."),
    ("Core plugin", "Semantic search", "Batch embedding generation + coverage stats",
     "Backfills embeddings for existing content.",
     "class-rest-api.php: /embeddings/batch, /stats", "U5", "Low", "No progress UI for long runs",
     "Keep - core", "Required for anyone importing an existing corpus."),
    ("Core plugin", "Semantic search", "Semantic search REST endpoint",
     "Direct similarity search API.",
     "class-rest-api.php: /search/semantic", "U2", "Low", "None",
     "Keep - core", "Extension point for add-on authors (U2)."),
    ("Core plugin", "Semantic search", "Corpus-wide RAG context mode",
     "Retrieves the most relevant content site-wide for the assistant.",
     "wp-copilot/includes/class-context-builder.php", "U5", "High", "Silently degrades without an OpenAI key",
     "Fix before launch", "Must clearly explain its requirement in the UI."),

    # ============================ CORE: IMPORT / EXPORT ============================
    ("Core plugin", "Import/export", "CSV export",
     "Exports the full hierarchy and items to a structured CSV.",
     "wp-copilot/includes/class-csv-exporter.php", "U1", "Low", "None",
     "Keep - core", "No-lock-in proof point. Name it in the marketing copy."),
    ("Core plugin", "Import/export", "CSV import (preview then commit)",
     "Two-phase dry-run import that reconstructs hierarchy from context paths.",
     "wp-copilot/includes/class-csv-importer.php", "U1", "Med", "None",
     "Keep - core", "The migration on-ramp from other tools."),
    ("Core plugin", "Import/export", "Raindrop.io bookmark import",
     "Scheduled/manual import of bookmarks as items.",
     "wp-copilot/includes/class-raindrop-importer.php", "U2", "Med", "Third API key; niche audience; cron surface",
     "Extract to add-on", "Perfect first proof that the add-on model works (U2). Out of core."),
    ("Core plugin", "Import/export", "Calendar .ics import",
     "Parses an uploaded .ics locally and shows events on the homepage.",
     "wp-copilot/includes/class-calendar-importer.php", "U2", "Med", "Homepage calendar card is half-finished",
     "Extract to add-on", "Weakly connected to the core loop. Second add-on candidate."),
    ("Core plugin", "Import/export", "Sample import CSV",
     "Example file shipped with the plugin.",
     "wp-copilot/sample-import.csv", "-", "Low", "None",
     "Keep - core", "Rewrite as a proper demo corpus for onboarding."),

    # ============================ CORE: STRUCTURE AUTOMATION ============================
    ("Core plugin", "Structure automation", "Page templates",
     "A parent page defines headings+items that new child pages inherit.",
     "wp-copilot/includes/class-page-template-manager.php", "U3", "High", "Item definitions are thinner than Section Manager's (see ROADMAP)",
     "Keep - flag off by default", "Strong concept, unfinished (Phase 2/3 unbuilt). Hide until it is coherent."),
    ("Core plugin", "Structure automation", "Page scheduler (cron-created pages)",
     "Creates recurring child pages on a schedule (e.g. weekly notes).",
     "wp-copilot/includes/class-page-scheduler.php", "U3", "High", "Autonomous writes contradict the human-in-the-loop guarantee",
     "Keep - flag off by default", "Directly conflicts with the 'AI never writes unattended' promise. Resolve the story first."),
    ("Core plugin", "Structure automation", "Section duplicate",
     "Clones a heading with its nested headings and items.",
     "wp-copilot/includes/class-section-manager.php", "U3", "Med", "Phase 1 of 3; the other phases are unbuilt",
     "Keep - flag off by default", "Ship when the template system it belongs to is finished."),
    ("Core plugin", "Structure automation", "Dynamic listings (saved filters)",
     "Saved filter queries rendered as sections on a page.",
     "class-rest-api.php: /dynamic-listings; theme dynamic-listing.php", "U3", "Med", "Overlaps search and fetch-posts",
     "Keep - core", "The best of the three overlapping query features - keep this one, hide the others."),
    ("Core plugin", "Structure automation", "Page notes metabox",
     "Freeform notes field on a page.",
     "wp-copilot/admin/class-page-notes-metabox.php", "-", "Low", "Overlaps page content and page mission",
     "Keep - flag off by default", "Three text fields on a page is two too many."),
    ("Core plugin", "Structure automation", "Saved prompt chips",
     "User-saved reusable prompts in the AI widget.",
     "class-rest-api.php: /prompts", "U5", "Low", "None",
     "Keep - core", "Cheap personalisation that raises repeat usage."),

    # ============================ CORE: ADMIN & PLATFORM ============================
    ("Core plugin", "Admin & platform", "Settings screen",
     "AI, embeddings and Raindrop configuration.",
     "wp-copilot/admin/class-settings.php", "U1", "Low", "Keys stored plainly in wp_options",
     "Fix before launch", "Document key handling; restrict the settings capability explicitly."),
    ("Core plugin", "Admin & platform", "Admin dashboard",
     "wp-admin overview screen.",
     "wp-copilot/admin/class-admin.php: render_dashboard()", "-", "Low", "Duplicates the front-end homepage dashboard",
     "Keep - flag off by default", "Two dashboards is clutter. The workspace one is the product (U7)."),
    ("Core plugin", "Admin & platform", "AI audit log screen",
     "Browsable history of every AI call and decision.",
     "wp-copilot/admin/class-admin.php: render_ai_log()", "U1", "Low", "No filtering, pagination or purge",
     "Fix before launch", "Named in readme.txt as a trust feature - it must hold up to scrutiny."),
    ("Core plugin", "Admin & platform", "CSV import/export screen",
     "wp-admin UI for the importer/exporter.",
     "wp-copilot/admin/class-admin.php: render_import_export()", "U1", "Low", "None",
     "Keep - core", "Fine where it is."),
    ("Core plugin", "Admin & platform", "REST API (~50 endpoints)",
     "Full API surface behind nonce/capability checks.",
     "wp-copilot/includes/class-rest-api.php (3,179 lines)", "U2", "High", "One endpoint uses __return_true; needs a capability audit",
     "Fix before launch", "Security-critical and the basis of the extensibility USP. Audit every permission_callback."),
    ("Core plugin", "Admin & platform", "Custom DB tables (5)",
     "ai_actions, embeddings, ai_conversations, ai_messages, edges.",
     "wp-copilot/work-copilot.php (activation)", "-", "Med", "No uninstall/cleanup routine; no dbDelta version guard",
     "Fix before launch", "WP.org reviewers look for uninstall.php. Add migration versioning."),
    ("Core plugin", "Admin & platform", "version-check.php probe",
     "Standalone file that reports the plugin version, referencing a localhost URL.",
     "wp-copilot/version-check.php", "-", "Low", "Directly-accessible PHP file outside the WP bootstrap - WP.org will reject this",
     "Delete", "Debug leftover and a security smell. Remove before submission."),

    # ============================ THEME ============================
    ("Theme", "Workspace UI", "Page template (main workspace view)",
     "The primary working surface: page, headings, items, listings.",
     "work-copilot-theme/page.php", "U7", "High", "None",
     "Keep - core", "This is the product."),
    ("Theme", "Workspace UI", "Sidebar navigation tree",
     "Collapsible page hierarchy nav.",
     "work-copilot-theme/sidebar.php", "U3, U7", "Med", "No search/filter within the tree at scale",
     "Keep - core", "How the folder-structure USP is experienced."),
    ("Theme", "Workspace UI", "Breadcrumbs",
     "Ancestor trail for pages, headings and items.",
     "work-copilot-theme/template-parts/breadcrumbs.php", "U3", "Low", "None",
     "Keep - core", "Essential orientation in a deep tree."),
    ("Theme", "Workspace UI", "Item row component",
     "Inline title edit, type/priority/status dropdowns, tags, contexts, subtasks, AI button.",
     "work-copilot-theme/template-parts/item-row.php", "U4", "High", "Dense - the biggest first-run comprehension risk",
     "Fix before launch", "Highest-leverage UX work available. Progressive disclosure of controls."),
    ("Theme", "Workspace UI", "Quick-add item form",
     "Add an item in context from anywhere on the page.",
     "work-copilot-theme/template-parts/quick-add-item.php", "U4", "Low", "None",
     "Keep - core", "Capture speed decides whether the habit forms."),
    ("Theme", "Workspace UI", "Heading sections",
     "Renders a heading with its items and its own quick-add.",
     "work-copilot-theme/template-parts/heading-section.php", "U3", "Med", "None",
     "Keep - core", "Core structural rendering."),
    ("Theme", "Workspace UI", "Drag-and-drop reorder + indent/outdent",
     "SortableJS vertical reorder and horizontal drag to nest.",
     "work-copilot-theme/assets/js/theme.js", "U4", "High", "Horizontal-drag nesting is unusual and error-prone on touch",
     "Fix before launch", "Verify on tablet/mobile or gate it to pointer devices."),
    ("Theme", "Workspace UI", "Keyboard indent/outdent (Ctrl+] / Ctrl+[)",
     "Outliner keyboard shortcuts while editing a title.",
     "work-copilot-theme/assets/js/theme.js", "U4", "Low", "Undiscoverable without documentation",
     "Keep - core", "Cheap; document it in the onboarding guide."),
    ("Theme", "Workspace UI", "Table of contents on page",
     "Jump list of headings within a page.",
     "work-copilot-theme/page.php", "U3", "Low", "None",
     "Keep - core", "Recently added; helps long pages."),
    ("Theme", "Workspace UI", "Markdown rendering in items",
     "Renders markdown in item content.",
     "work-copilot-theme/assets/js/theme.js", "U4", "Med", "Sanitisation of AI-produced markdown needs review",
     "Fix before launch", "XSS surface - confirm output escaping before launch."),
    ("Theme", "Workspace UI", "Floating AI widget",
     "The assistant panel present on every page.",
     "work-copilot-theme/template-parts/ai-widget.php (+1,731 lines JS)", "U3, U5", "High", "Action menu is long enough to overwhelm",
     "Fix before launch", "Cutting the flagged-off actions above is largely this file's cleanup."),
    ("Theme", "Homepage", "Dashboard panel (tasks, this week, overdue)",
     "Default homepage view of what needs attention.",
     "work-copilot-theme/index.php", "U7", "Med", "Card set has grown organically",
     "Keep - core", "First screen after login - decides the daily-use habit."),
    ("Theme", "Homepage", "Homepage AI chat panel",
     "Corpus-level chat from the homepage.",
     "work-copilot-theme/index.php", "U3", "Med", "None",
     "Keep - core", "Recently added; the natural 'ask anything' entry point."),
    ("Theme", "Homepage", "Structure tree panel",
     "Whole-corpus tree browser on the homepage.",
     "work-copilot-theme/index.php", "U3", "Low", "Duplicates the sidebar",
     "Keep - flag off by default", "Redundant with the sidebar nav."),
    ("Theme", "Homepage", "Activity panel + AI activity summary",
     "Recent changes grouped by page, with an AI summary button.",
     "work-copilot-theme/index.php", "U6", "Med", "None",
     "Keep - core", "Makes the learning loop visible."),
    ("Theme", "Homepage", "Calendar card (.ics upload + week view)",
     "Upload a calendar file and see the week.",
     "work-copilot-theme/index.php", "-", "Med", "Half-built; belongs with the calendar importer",
     "Extract to add-on", "Moves out with the calendar import add-on."),
    ("Theme", "Homepage", "Upcoming scheduled pages card",
     "Preview of pages the scheduler will create.",
     "work-copilot-theme/index.php", "-", "Low", "Ships or hides with the page scheduler",
     "Keep - flag off by default", "Tied to a flagged-off feature."),
    ("Theme", "Navigation", "Search page (items-only)",
     "Intercepts WP search to return items rather than pages.",
     "work-copilot-theme/search.php", "U4", "Low", "Keyword only; no filters",
     "Keep - core", "Baseline expectation."),
    ("Theme", "Navigation", "Tag archive page",
     "Browse all items with a given tag.",
     "work-copilot-theme/tag.php", "U4", "Low", "None",
     "Keep - core", "Part of the tagging story."),
    ("Theme", "Navigation", "Single item view",
     "Full view of one item with its contexts and children.",
     "work-copilot-theme/single.php", "U4", "Low", "None",
     "Keep - core", "Needed for canonical references."),
    ("Theme", "Platform", "REST auth lockdown filter",
     "Blocks all anonymous REST access site-wide.",
     "work-copilot-theme/functions.php:754", "U1", "Low", "A theme changing site-wide REST behaviour is surprising and will break other plugins",
     "Fix before launch", "Correct instinct, wrong layer. Move into the plugin and scope it to this namespace."),
    ("Theme", "Platform", "Theme CSS (3,612 lines) + workspace UI + AI widget CSS",
     "All visual styling.",
     "work-copilot-theme/assets/css/", "U7", "High", "Grown organically; a revert commit suggests an abandoned redesign",
     "Fix before launch", "Marketing screenshots come from here. Budget explicit design time."),
    ("Theme", "Platform", "Theme is required for the plugin to be usable",
     "Core workspace UI lives in the theme, not the plugin.",
     "work-copilot-theme/", "U2, U7", "High", "WP.org distributes plugins and themes separately - this is a structural launch blocker",
     "Fix before launch", "DECIDE FIRST: bundle the UI into the plugin, or ship a plugin+theme pair. Everything else depends on this."),

    # ============================ ADD-ONS ============================
    ("Voice add-on", "Capture", "Push-to-talk voice capture",
     "Hold the backtick key to dictate a new item via the Web Speech API.",
     "wp-copilot-voice/", "U2", "Low", "HTTPS only; Chrome-family browsers only; key binding conflicts with code blocks",
     "Extract to add-on", "Delightful demo material. Ship as a separate free add-on to prove U2."),
    ("Graph add-on", "Relationships", "Triple store (subject-predicate-object)",
     "Custom edges table linking pages, items and headings.",
     "wcp-graph/includes/class-graph-repository.php", "U2, U4", "High", "Concept overhead on top of an already rich model",
     "Extract to add-on", "Genuinely differentiated, but too much for a first-run user. Sell as an add-on."),
    ("Graph add-on", "Relationships", "Connections panel",
     "Relationship chips on pages and items.",
     "wcp-graph/includes/class-connections-panel.php", "U2", "Med", "None",
     "Extract to add-on", "Moves with the graph plugin."),
    ("Graph add-on", "Relationships", "Predicates registry",
     "Defines the allowed relationship verbs.",
     "wcp-graph/includes/class-predicates.php", "U2", "Low", "None",
     "Extract to add-on", "Moves with the graph plugin."),
    ("Graph add-on", "Relationships", "Semantic tables",
     "Entity/predicate grid view for bulk relationship editing.",
     "wcp-graph/includes/class-semantic-tables.php", "U2", "High", "Most complex UI in the codebase for the narrowest use case",
     "Move to backup repo", "Park it. Revisit only if the graph add-on finds an audience."),
    ("Delegation add-on", "Agents", "Delegation manager + work packets",
     "Hands an item to an external agent and tracks its execution.",
     "wcp-delegation/includes/class-delegation-manager.php", "U2, U6", "High", "Built around one specific private agent ('Hermes')",
     "Move to backup repo", "Not generalisable as-is. Rebuild later as a documented agent-integration API."),
    ("Delegation add-on", "Agents", "Telegram notifications",
     "Pings a Telegram chat on delegation events.",
     "wcp-delegation/includes/class-delegation-manager.php", "-", "Med", "Fourth external service",
     "Move to backup repo", "Moves with delegation."),
    ("Delegation add-on", "Agents", "Clarification loop + artifact upload",
     "Agent asks questions and returns files against an item.",
     "wcp-delegation/includes/class-delegation-manager.php", "U6", "High", "None",
     "Move to backup repo", "Best idea in the delegation plugin - preserve the design notes for the future API."),
    ("Delegation add-on", "Platform", "Caddy basic-auth gate shim",
     "Works around a reverse proxy on the author's own server; hardcodes the username 'michael'.",
     "wcp-delegation/wcp-delegation.php:33", "-", "Low", "Personal-infrastructure code in a public plugin - immediate WP.org rejection",
     "Delete", "Must not ship under any circumstances."),

    # ============================ SEPARATE VERTICALS ============================
    ("Wiretap (separate product)", "Vertical", "KOL tracking, tweet ingest, prefilter",
     "Monitors investment influencers on X and detects actionable calls.",
     "wcp-wiretap/ (~3,000 lines)", "-", "High", "Unrelated vertical; financial-advice surface; X API dependency",
     "Move to backup repo", "A different product with a different audience. Excluding it clarifies the story."),
    ("Wiretap (separate product)", "Vertical", "Trade plans, earliness scoring, digest",
     "Scores signal timeliness and proposes conditional trade plans.",
     "wcp-wiretap/includes/", "-", "High", "Financial-recommendation liability; WP.org guideline risk",
     "Move to backup repo", "Keeping this in the repo actively harms the launch narrative."),
    ("Wiretap (separate product)", "Vertical", "KOL discovery, themes, runs & budget",
     "Finds new accounts to follow and tracks AI spend.",
     "wcp-wiretap/includes/", "-", "High", "Unrelated vertical",
     "Move to backup repo", "Moves with wiretap. (Note: the budget-tracking idea is worth porting to core.)"),
    ("OpenBiografy (separate product)", "Vertical", "Fact extraction and reconciliation",
     "Ingests documents about a person and extracts atomic facts as proposals.",
     "wcp-openbiografy/ (~2,500 lines)", "-", "High", "Unrelated vertical; PII/biography-of-real-people risk",
     "Move to backup repo", "Different product. The extract-and-reconcile pattern is worth porting to core later."),
    ("OpenBiografy (separate product)", "Vertical", "Timeline, chapters, EDTF dates, export",
     "Reconciles facts into a life timeline and drafts narrative chapters.",
     "wcp-openbiografy/includes/", "-", "High", "Unrelated vertical",
     "Move to backup repo", "Moves with openbiografy."),

    # ============================ REPO TOOLING ============================
    ("Theme", "Platform", "Third-party JS loaded from a CDN",
     "SortableJS and marked.js are enqueued from cdn.jsdelivr.net.",
     "work-copilot-theme/functions.php:61,66", "-", "Low", "WP.org forbids loading assets from third-party CDNs - hard rejection, and it leaks visitor IPs",
     "Fix before launch", "Bundle both libraries locally with their licences. Mechanical fix, absolute blocker."),
    ("Core plugin", "Admin & platform", "DB schema versioning",
     "wcp_db_version option is written, but dbDelta only runs on the activation hook.",
     "wp-copilot/work-copilot.php:71,243", "-", "Med", "A plugin update never runs migrations - existing installs silently keep the old schema",
     "Fix before launch", "Add an admin_init version check that re-runs dbDelta when the stored version is behind."),

    ("Repo tooling", "Ops", "deploy.sh",
     "Commits, pushes and SSHes into a specific server to git pull.",
     "deploy.sh", "-", "Low", "Hardcoded production server IP in a public repo",
     "Delete", "Remove before the repo goes public. Replace with a documented release process."),
    ("Repo tooling", "Ops", "Bookmark categoriser script + CSV data",
     "Python script and personal tweet/bookmark CSVs.",
     "scripts/", "-", "Low", "Contains the author's personal data and stray lock files",
     "Move to backup repo", "Personal data must not ship in a public repo."),
    ("Repo tooling", "Docs", "readme.txt (WP.org format)",
     "Plugin directory listing with an External Services disclosure.",
     "wp-copilot/readme.txt", "U1", "Low", "Placeholder author/contributor fields; version claims need re-verifying",
     "Fix before launch", "Already strong - the External Services section is exactly what reviewers want. Fill in the placeholders."),
    ("Repo tooling", "Docs", "DOCUMENTATION.md / QUICK-START.md / prd.md / ROADMAP.md",
     "Internal product and design documentation.",
     "repo root", "-", "Low", "Written for the author, not for users",
     "Keep - core", "Excellent raw material for the public docs site. Rewrite for an external reader."),
    ("Repo tooling", "Docs", "PRD-COMPOUNDING-LOOPS.md (48KB)",
     "Long-form internal strategy document.",
     "PRD-COMPOUNDING-LOOPS.md", "-", "Low", "Internal strategy in a soon-to-be-public repo",
     "Move to backup repo", "Keep it private."),
    ("Repo tooling", "Docs", "AGENTS.md / claude.md / agent-os standards",
     "Coding standards and agent instructions for this repo.",
     "AGENTS.md, claude.md, agent-os/", "-", "Low", "None",
     "Keep - core", "Useful signal to open-source contributors that the repo has standards."),
]

HEADERS = [
    ("ID", 6),
    ("Bundle", 20),
    ("Area", 20),
    ("Feature", 34),
    ("What it does", 46),
    ("Primary code location", 38),
    ("USP served", 12),
    ("Effort to keep", 12),
    ("Launch risk / known issue", 44),
    ("Proposed action", 22),
    ("Rationale", 50),
    ("YOUR DECISION", 20),
    ("Owner", 12),
    ("Notes", 30),
]

USPS = [
    ("U1", "Open source, bring your own API key - no vendor account, no lock-in"),
    ("U2", "Extensible via add-on plugins, some of them ours"),
    ("U3", "Organised AI chats - a page structure for your areas and projects, chats live inside pages"),
    ("U4", "Atomic items - every action or statement is its own taggable, multi-filed object with canonical references"),
    ("U5", "Page-scoped context plus a configurable mission for the install and for each page"),
    ("U6", "The assistant co-authors into the corpus, creating a learning loop against your own protocols"),
    ("U7", "No terminal, no code, no repo to manage - it all happens in a website"),
]


def style_header(ws, row_idx, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = H_FILL
        cell.font = H_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BOX


def build():
    wb = Workbook()

    # ------------------------------------------------------- Sheet 1: README
    ws = wb.active
    ws.title = "Read me"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 100

    ws["B2"] = "Work Copilot - Feature Inventory & Launch Scope Decisions"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = "Built from a read of the codebase at commit a4cf254 (main). Every row cites where the code lives."
    ws["B3"].font = SUB_FONT

    r = 5
    ws.cell(row=r, column=2, value="How to use this").font = BODY_B
    r += 1
    for line in [
        "1. Work down the 'Feature inventory' tab. The 'Proposed action' column is a recommendation, not a decision.",
        "2. Put your call in the yellow 'YOUR DECISION' column (dropdown). It is the only column you need to fill in.",
        "3. The 'Scope summary' tab counts your decisions live as you make them.",
        "4. Anything marked 'Move to backup repo' should be cut to a separate private repository before the public repo goes live.",
    ]:
        ws.cell(row=r, column=3, value=line).font = BODY
        ws.cell(row=r, column=3).alignment = WRAP
        r += 1

    r += 1
    ws.cell(row=r, column=2, value="Action key").font = BODY_B
    r += 1
    for label, meaning in [
        ("Keep - core", "Ships enabled in v1. Part of the core promise."),
        ("Keep - flag off by default", "Code stays, hidden behind a feature flag. Available to power users, invisible on first run."),
        ("Fix before launch", "Ships, but has a defect or gap that blocks launch as it stands."),
        ("Extract to add-on", "Leaves core, ships as a separate plugin. Proves the extensibility USP."),
        ("Move to backup repo", "Cut from the public repository entirely; preserved privately."),
        ("Delete", "Remove outright. Dead, duplicated, or unsafe to publish."),
    ]:
        c = ws.cell(row=r, column=2, value=label)
        c.font = BODY_B
        c.fill = REC_FILLS[label]
        c.border = BOX
        m = ws.cell(row=r, column=3, value=meaning)
        m.font = BODY
        m.alignment = WRAP
        r += 1

    r += 1
    ws.cell(row=r, column=2, value="USP reference").font = BODY_B
    r += 1
    for code, text in USPS:
        ws.cell(row=r, column=2, value=code).font = BODY_B
        ws.cell(row=r, column=3, value=text).font = BODY
        ws.cell(row=r, column=3).alignment = WRAP
        r += 1

    r += 1
    ws.cell(row=r, column=2, value="Assumptions").font = BODY_B
    r += 1
    for line in [
        "Feature list derived by reading the source, not from a product backlog - a feature with no code is not listed here.",
        "'Effort to keep' is a judgement of ongoing maintenance burden (code size, external dependencies, coupling), not build cost.",
        "'Launch risk' flags issues visible from reading the code. It is not a substitute for a QA pass or a security audit.",
        "Line counts and file paths are accurate as of the commit above and will drift as work proceeds.",
    ]:
        ws.cell(row=r, column=3, value=line).font = BODY
        ws.cell(row=r, column=3).alignment = WRAP
        r += 1

    # ---------------------------------------------- Sheet 2: Feature inventory
    fi = wb.create_sheet("Feature inventory")
    fi.sheet_view.showGridLines = False

    for i, (h, w) in enumerate(HEADERS, start=1):
        fi.cell(row=1, column=i, value=h)
        fi.column_dimensions[get_column_letter(i)].width = w
    style_header(fi, 1, len(HEADERS))
    fi.row_dimensions[1].height = 30
    fi.freeze_panes = "D2"

    for idx, row in enumerate(ROWS, start=1):
        r = idx + 1
        bundle, area, feature, what, loc, usp, effort, risk, rec, rationale = row
        values = [idx, bundle, area, feature, what, loc, usp, effort, risk, rec, rationale, None, None, None]
        for c, v in enumerate(values, start=1):
            cell = fi.cell(row=r, column=c, value=v)
            cell.font = BODY
            cell.alignment = WRAP
            cell.border = BOX
        fi.cell(row=r, column=4).font = BODY_B
        fi.cell(row=r, column=10).fill = REC_FILLS[rec]
        for c in (12, 13, 14):
            fi.cell(row=r, column=c).font = INPUT_FONT
            fi.cell(row=r, column=c).fill = INPUT_FILL

    last = len(ROWS) + 1

    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(RECOMMENDATIONS) + '"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.prompt = "Pick your call for this feature."
    dv.promptTitle = "Launch decision"
    fi.add_data_validation(dv)
    dv.add(f"L2:L{last}")

    fi.auto_filter.ref = f"A1:N{last}"

    # ------------------------------------------------- Sheet 3: Scope summary
    su = wb.create_sheet("Scope summary")
    su.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", (3, 30, 16, 16, 52)):
        su.column_dimensions[col].width = w

    su["B2"] = "Scope summary"
    su["B2"].font = TITLE_FONT
    su["B3"] = "Counts recalculate from the Feature inventory tab. 'Yours' stays at zero until you fill in the decision column."
    su["B3"].font = SUB_FONT
    su["B4"] = ("These cells are live formulas. They compute the moment the file opens in Excel, "
                "LibreOffice, Numbers or Google Sheets - if a preview shows them blank, that preview "
                "is not calculating, not a broken sheet.")
    su["B4"].font = SUB_FONT
    su.merge_cells("B4:E4")
    su["B4"].alignment = WRAP
    su.row_dimensions[4].height = 28

    su["B5"] = "By proposed action"
    su["B5"].font = BODY_B
    for i, h in enumerate(["Action", "Proposed", "Yours", "What it means for the launch"], start=2):
        c = su.cell(row=6, column=i, value=h)
        c.fill = H_FILL
        c.font = H_FONT
        c.border = BOX

    meanings = {
        "Keep - core": "The v1 surface area a new user sees.",
        "Keep - flag off by default": "Code retained, hidden. Zero onboarding cost.",
        "Fix before launch": "Your pre-launch engineering backlog.",
        "Extract to add-on": "Separate plugins - and the proof that the platform is extensible.",
        "Move to backup repo": "Cut from the public repo before it goes live.",
        "Delete": "Removed outright.",
    }
    row = 7
    for rec in RECOMMENDATIONS:
        su.cell(row=row, column=2, value=rec).font = BODY_B
        su.cell(row=row, column=2).fill = REC_FILLS[rec]
        su.cell(row=row, column=3,
                value=f"=COUNTIF('Feature inventory'!$J$2:$J${last},$B{row})").font = BODY
        su.cell(row=row, column=4,
                value=f"=COUNTIF('Feature inventory'!$L$2:$L${last},$B{row})").font = BODY
        su.cell(row=row, column=5, value=meanings[rec]).font = BODY
        for c in range(2, 6):
            su.cell(row=row, column=c).border = BOX
            su.cell(row=row, column=c).alignment = WRAP
        row += 1

    su.cell(row=row, column=2, value="Total features catalogued").font = BODY_B
    su.cell(row=row, column=3, value=f"=COUNTA('Feature inventory'!$D$2:$D${last})").font = BODY_B
    su.cell(row=row, column=4, value=f"=COUNTA('Feature inventory'!$L$2:$L${last})").font = BODY_B
    su.cell(row=row, column=5, value="Column D of the inventory; 'Yours' counts decisions made so far.").font = BODY
    for c in range(2, 6):
        su.cell(row=row, column=c).fill = GROUP_FILL
        su.cell(row=row, column=c).border = BOX
        su.cell(row=row, column=c).alignment = WRAP

    bundles = []
    for r_ in ROWS:
        if r_[0] not in bundles:
            bundles.append(r_[0])

    row += 3
    su.cell(row=row, column=2, value="By bundle").font = BODY_B
    row += 1
    for i, h in enumerate(["Bundle", "Features", "Keep - core",
                           "Share proposed as core"], start=2):
        c = su.cell(row=row, column=i, value=h)
        c.fill = H_FILL
        c.font = H_FONT
        c.border = BOX
    row += 1
    for b in bundles:
        su.cell(row=row, column=2, value=b).font = BODY
        su.cell(row=row, column=3,
                value=f"=COUNTIF('Feature inventory'!$B$2:$B${last},$B{row})").font = BODY
        su.cell(row=row, column=4,
                value=f"=COUNTIFS('Feature inventory'!$B$2:$B${last},$B{row},'Feature inventory'!$J$2:$J${last},\"Keep - core\")").font = BODY
        su.cell(row=row, column=5,
                value=f"=IF($C{row}=0,0,$D{row}/$C{row})").font = BODY
        su.cell(row=row, column=5).number_format = "0.0%;(0.0%);-"
        for c in range(2, 6):
            su.cell(row=row, column=c).border = BOX
            su.cell(row=row, column=c).alignment = WRAP
        row += 1

    row += 2
    su.cell(row=row, column=2, value="Reading of the numbers").font = BODY_B
    row += 1
    for line in [
        "Roughly a third of everything catalogued is a separate product (Wiretap, OpenBiografy) or personal infrastructure. Cutting it is the single biggest clarity win available, and costs nothing the launch needs.",
        "The 'Fix before launch' list is the real launch gate. The largest item on it is structural: core workspace UI currently lives in the theme, not the plugin.",
        "Feature-flagging rather than deleting keeps optionality. Nothing marked 'flag off' is bad - it is unfinished, redundant with a better feature, or too advanced for a first run.",
    ]:
        su.cell(row=row, column=2, value=line).font = BODY
        su.cell(row=row, column=2).alignment = WRAP
        su.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        su.row_dimensions[row].height = 30
        row += 2

    # ------------------------------------------- Sheet 4: Launch blocker list
    lb = wb.create_sheet("Launch blockers")
    lb.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", (3, 8, 40, 60, 14, 18)):
        lb.column_dimensions[col].width = w

    lb["B2"] = "Launch blockers - everything marked 'Fix before launch' or 'Delete'"
    lb["B2"].font = TITLE_FONT
    lb["B3"] = "Ordered by what would hurt most if it shipped as-is. Sequence, not estimates - sizing is a conversation with whoever builds it."
    lb["B3"].font = SUB_FONT

    for i, h in enumerate(["#", "Blocker", "Why it blocks", "Severity", "Your owner"], start=2):
        c = lb.cell(row=5, column=i, value=h)
        c.fill = H_FILL
        c.font = H_FONT
        c.border = BOX

    blockers = [
        ("Theme vs plugin split", "Core workspace UI lives in the theme. WP.org distributes plugins and themes separately, so 'install the plugin' does not currently give anyone a working product. Decide: bundle the UI into the plugin, or ship an explicit plugin+theme pair.", "Critical"),
        ("Caddy basic-auth shim hardcoding 'michael'", "Personal server infrastructure in shippable code. Immediate rejection, and an embarrassment in a public repo.", "Critical"),
        ("deploy.sh with a production server IP", "Public repo would expose the author's server address.", "Critical"),
        ("version-check.php", "Directly-accessible PHP outside the WordPress bootstrap. WP.org rejects this pattern.", "Critical"),
        ("Personal data in scripts/", "Author's own tweet and bookmark CSVs, plus editor lock files, committed to the repo.", "Critical"),
        ("REST permission_callback audit", "At least one endpoint uses __return_true. Every one of ~50 endpoints needs an explicit capability check.", "Critical"),
        ("SortableJS + marked.js loaded from jsDelivr", "WP.org forbids enqueuing assets from third-party CDNs. Also leaks every visitor's IP to a third party. Bundle both locally with their licence files.", "Critical"),
        ("dbDelta only runs on activation", "wcp_db_version is stored but never compared on update, so a plugin upgrade never migrates an existing install's schema. This breaks people silently, and only after they have data worth losing.", "High"),
        ("No uninstall.php / no DB migration versioning", "Five custom tables are created with no cleanup path and no schema version guard. Reviewers check for this.", "High"),
        ("REST lockdown filter lives in the theme", "A theme silently changing site-wide REST behaviour will break other plugins. Move to the plugin, scope to this namespace.", "High"),
        ("AI audit log has no retention or purge", "Table grows unbounded and stores full prompts. It is named in readme.txt as a privacy feature, so it has to hold up.", "High"),
        ("Bulk 'edit items' JSON parsing", "Already patched twice (see changelog). Highest blast radius of any AI action - it rewrites many items at once.", "High"),
        ("Markdown rendering sanitisation", "AI-produced markdown rendered to the page. Confirm escaping before anyone else runs this.", "High"),
        ("Rewrite page content has no undo", "Destructive and irreversible from the UI. Snapshot to a WP revision first.", "High"),
        ("Anthropic client error handling", "No retry, no backoff, thin error surfacing. BYO-key is the headline USP - a bad key must produce a clear message, not a silent failure.", "High"),
        ("Hardcoded model IDs in the widget template", "Model list will rot with each release. Move to a filterable config array.", "Medium"),
        ("Embeddings generated synchronously on save", "Save latency and API cost on every edit. Queue it.", "Medium"),
        ("Silent RAG degradation without an OpenAI key", "'Entire corpus' quietly does nothing. Must explain itself in the UI.", "Medium"),
        ("Page mission only editable in wp-admin", "A headline USP is hidden in the admin area, contradicting the 'it all happens in a website' pitch.", "Medium"),
        ("Item row density", "The single biggest first-run comprehension risk in the product. Needs progressive disclosure.", "Medium"),
        ("Horizontal-drag nesting on touch devices", "Unusual interaction, likely broken on tablets. Verify or gate to pointer devices.", "Medium"),
        ("readme.txt placeholder fields", "Author, contributor and URI fields still say 'Your Name' / 'yoursite.com'.", "Medium"),
        ("CSS state after the reverted redesign", "A redesign was committed then reverted. Marketing screenshots come from here - budget explicit design time.", "Medium"),
    ]

    row = 6
    for i, (name, why, sev) in enumerate(blockers, start=1):
        lb.cell(row=row, column=2, value=i).font = BODY
        lb.cell(row=row, column=3, value=name).font = BODY_B
        lb.cell(row=row, column=4, value=why).font = BODY
        sc = lb.cell(row=row, column=5, value=sev)
        sc.font = BODY_B
        sc.fill = {"Critical": REC_FILLS["Delete"],
                   "High": REC_FILLS["Fix before launch"],
                   "Medium": REC_FILLS["Keep - flag off by default"]}[sev]
        oc = lb.cell(row=row, column=6)
        oc.font = INPUT_FONT
        oc.fill = INPUT_FILL
        for c in range(2, 7):
            lb.cell(row=row, column=c).border = BOX
            lb.cell(row=row, column=c).alignment = WRAP
        row += 1

    row += 1
    lb.cell(row=row, column=3, value="Severity key: Critical = cannot publish the repo or submit to WP.org. High = will produce data loss, a security report, or a bad first review. Medium = will cost adoption but is survivable.").font = SUB_FONT
    lb.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    lb.cell(row=row, column=3).alignment = WRAP

    wb.save("/home/user/work-copilot/gtm/work-copilot-feature-inventory.xlsx")
    print(f"Wrote {len(ROWS)} feature rows, {len(blockers)} blockers.")


if __name__ == "__main__":
    build()
